# @Time     :2020.6.17
# @Author   :wanggaojian
# @File     :lesson

from openpyxl import load_workbook
def read_data(file_name,sheet_name):
    wb=load_workbook(file_name)
    sheet=wb[sheet_name]
    all_case=[]
    for i in range(2,sheet.max_row+1):            #range(2,sheet.max_row+lemon)=range(2,12+lemon)=range(2,13)  读取2 3 4 5 6 7 8 9 10 11 12
        case=[]
        for j in range(1,sheet.max_column-1):     #range(lemon,sheet.max_column-lemon)=range(lemon,9-lemon)=range(lemon,8)  读取1 2 3 4 5 6 7
            case.append(sheet.cell(row=i,column=j).value)
        all_case.append(case)
    return all_case

def write_data(file_name,sheet_name,row,column,value):
    wb = load_workbook(file_name)
    sheet = wb[sheet_name]
    sheet.cell(row=row, column=column).value = value
    wb.save(file_name)

if __name__ == '__main__':
    all_case=read_data('test_case6.xlsx','Sheet1')
    print("所有的测试数据：{}".format(all_case))